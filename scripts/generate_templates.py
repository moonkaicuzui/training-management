#!/usr/bin/env python3
"""
Generate Excel templates for Q-TRAIN system master data registration.
Creates one multi-sheet workbook with all required data templates.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

# ─── Style Constants ──────────────────────────────────────────────────────────

HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
EXAMPLE_FILL = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
REQUIRED_FILL = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
OPTIONAL_FILL = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid')
NOTE_FONT = Font(name='Calibri', italic=True, size=10, color='6B7280')
THIN_BORDER = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


def style_header(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_row(ws, row, num_cols, fill=None):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.alignment = LEFT
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def add_legend(ws, start_row, num_cols):
    """Add color legend at bottom."""
    r = start_row + 2
    ws.cell(row=r, column=1, value='LEGEND:').font = Font(bold=True, size=10)
    r += 1
    c1 = ws.cell(row=r, column=1, value='Example row (yellow)')
    c1.fill = EXAMPLE_FILL
    c1.font = NOTE_FONT
    r += 1
    c2 = ws.cell(row=r, column=1, value='* = Required field')
    c2.font = Font(bold=True, size=10, color='DC2626')
    r += 1
    ws.cell(row=r, column=1, value='Please fill in your data starting from row 4.').font = NOTE_FONT
    ws.cell(row=r + 1, column=1, value='Do not modify header rows (rows 1-2).').font = NOTE_FONT


# ═══════════════════════════════════════════════════════════════════════════════
# 1. EMPLOYEES SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def create_employees_sheet(wb):
    ws = wb.create_sheet('1. Employees')

    headers = [
        ('employee_id *', 15),
        ('employee_name *', 25),
        ('department *', 15),
        ('position *', 20),
        ('building *', 20),
        ('line *', 10),
        ('hire_date *\n(YYYY-MM-DD)', 15),
        ('status *', 12),
    ]

    notes = [
        'Unique ID\n(e.g. HV0001)',
        'Full name',
        'See dropdown',
        'See dropdown',
        'See dropdown',
        'Line number\nor name',
        'Format:\nYYYY-MM-DD',
        'ACTIVE or\nINACTIVE',
    ]

    # Header row
    for i, (h, w) in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws, 1, len(headers))

    # Notes row
    for i, n in enumerate(notes, 1):
        ws.cell(row=2, column=i, value=n).font = NOTE_FONT
        ws.cell(row=2, column=i).alignment = CENTER
        ws.cell(row=2, column=i).border = THIN_BORDER

    # Example rows
    examples = [
        ['HV0001', 'Nguyen Van A', 'ASSEMBLY', 'QIP_TQC', 'BUILDING_A', 'L1', '2023-01-15', 'ACTIVE'],
        ['HV0002', 'Tran Thi B', 'STITCHING', 'PRO_WORKER', 'BUILDING_B', 'L3', '2022-06-01', 'ACTIVE'],
        ['HV0003', 'Le Van C', 'CUTTING', 'QIP_LINE_LEADER', 'BUILDING_C', 'L2', '2021-03-20', 'ACTIVE'],
        ['HV0004', 'Pham Thi D', 'QA', 'QIP_QA', 'BUILDING_QA_OFFICE', 'QA1', '2024-01-10', 'ACTIVE'],
        ['HV0005', 'Vo Van E', 'OSC', 'PRO_LINE_LEADER', 'BUILDING_OSC_A', 'O1', '2020-08-05', 'ACTIVE'],
    ]

    for r, ex in enumerate(examples, 3):
        for c, val in enumerate(ex, 1):
            ws.cell(row=r, column=c, value=val)
        style_row(ws, r, len(headers), EXAMPLE_FILL)

    # Data validation
    dept_list = '"ASSEMBLY,STITCHING,CUTTING,OSC,MTL,BOTTOM,AQL,REPACKING,QA,OFFICE,NEW"'
    dv_dept = DataValidation(type='list', formula1=dept_list, allow_blank=False)
    dv_dept.error = 'Please select a valid department'
    dv_dept.errorTitle = 'Invalid Department'
    ws.add_data_validation(dv_dept)
    dv_dept.add(f'C3:C500')

    pos_list = '"QIP_TQC,QIP_RQC,QIP_CFA,QIP_QA,QIP_LINE_LEADER,QIP_GROUP_LEADER,QIP_SUPERVISOR,QIP_MANAGER_PLUS,QIP_OFFICE,QIP_NEW_MEMBER,PRO_WORKER,PRO_LINE_LEADER,PRO_GROUP_PLUS"'
    dv_pos = DataValidation(type='list', formula1=pos_list, allow_blank=False)
    ws.add_data_validation(dv_pos)
    dv_pos.add(f'D3:D500')

    bldg_list = '"BUILDING_A,BUILDING_A1,BUILDING_A2,BUILDING_B,BUILDING_B1,BUILDING_B2,BUILDING_B3,BUILDING_C,BUILDING_D,BUILDING_E1,BUILDING_E2,BUILDING_EZ_HAPPO,BUILDING_FG_WH,BUILDING_INHOUSE_EZ,BUILDING_INHOUSE_PRINTING,BUILDING_MTL_WH,BUILDING_OSC_A,BUILDING_QA_OFFICE,BUILDING_QIP_OFFICE"'
    dv_bldg = DataValidation(type='list', formula1=bldg_list, allow_blank=False)
    ws.add_data_validation(dv_bldg)
    dv_bldg.add(f'E3:E500')

    dv_status = DataValidation(type='list', formula1='"ACTIVE,INACTIVE"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add(f'H3:H500')

    add_legend(ws, 3 + len(examples), len(headers))
    ws.freeze_panes = 'A3'


# ═══════════════════════════════════════════════════════════════════════════════
# 2. TRAINERS SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def create_trainers_sheet(wb):
    ws = wb.create_sheet('1. Trainers')

    headers = [
        ('trainer_name *', 25),
        ('trainer_type *', 15),
        ('department', 15),
        ('company', 20),
        ('email', 25),
        ('phone', 15),
        ('specializations\n(comma-separated)', 30),
        ('certifications\n(comma-separated)', 30),
        ('notes', 30),
    ]

    notes = [
        'Full name',
        'INTERNAL or\nEXTERNAL',
        'If INTERNAL:\ndepartment',
        'If EXTERNAL:\ncompany name',
        'Email address',
        'Phone number',
        'Program codes\nthey can teach',
        'Certifications\nheld',
        'Additional notes',
    ]

    for i, (h, w) in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws, 1, len(headers))

    for i, n in enumerate(notes, 1):
        ws.cell(row=2, column=i, value=n).font = NOTE_FONT
        ws.cell(row=2, column=i).alignment = CENTER
        ws.cell(row=2, column=i).border = THIN_BORDER

    examples = [
        ['Kim Anh', 'INTERNAL', 'QA', '', 'kimanh@hsvina.com', '0901234567', 'QIP-001, QIP-002, QIP-003', 'ISO 9001 Auditor', 'Senior QIP trainer'],
        ['Nguyen Van Thanh', 'INTERNAL', 'ASSEMBLY', '', 'thanh.nv@hsvina.com', '0912345678', 'PRO-001, PRO-002', '', 'Assembly line specialist'],
        ['Dr. Park Jihun', 'EXTERNAL', '', 'Korea Quality Institute', 'parkjh@kqi.co.kr', '+82-10-1234-5678', 'QIP-010, QIP-011', 'PhD Quality Engineering, 6 Sigma BB', 'Annual visit trainer'],
    ]

    for r, ex in enumerate(examples, 3):
        for c, val in enumerate(ex, 1):
            ws.cell(row=r, column=c, value=val)
        style_row(ws, r, len(headers), EXAMPLE_FILL)

    dv_type = DataValidation(type='list', formula1='"INTERNAL,EXTERNAL"', allow_blank=False)
    ws.add_data_validation(dv_type)
    dv_type.add('B3:B500')

    add_legend(ws, 3 + len(examples), len(headers))
    ws.freeze_panes = 'A3'


# ═══════════════════════════════════════════════════════════════════════════════
# 3. COMPETENCIES SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def create_competencies_sheet(wb):
    ws = wb.create_sheet('2. Competencies')

    headers = [
        ('competency_code *', 18),
        ('name (EN) *', 25),
        ('name (VI) *', 25),
        ('name (KR) *', 25),
        ('category *', 18),
        ('description *', 40),
        ('is_core *', 10),
        ('related_programs\n(comma-separated)', 30),
    ]

    notes = [
        'Unique code\n(e.g. COMP-001)',
        'English name',
        'Vietnamese name',
        'Korean name',
        'See dropdown',
        'Detailed description\nof this competency',
        'YES or NO',
        'Related program\ncodes',
    ]

    for i, (h, w) in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws, 1, len(headers))

    for i, n in enumerate(notes, 1):
        ws.cell(row=2, column=i, value=n).font = NOTE_FONT
        ws.cell(row=2, column=i).alignment = CENTER
        ws.cell(row=2, column=i).border = THIN_BORDER

    examples = [
        ['COMP-001', 'Visual Inspection', 'Kiểm tra trực quan', '외관 검사', 'QUALITY', 'Ability to identify product defects through visual examination, including color, shape, and surface quality assessment', 'YES', 'QIP-001, QIP-005'],
        ['COMP-002', 'Stitching Quality Control', 'Kiểm soát chất lượng may', '봉제 품질 관리', 'TECHNICAL', 'Knowledge and skills for monitoring stitching quality parameters, including stitch count, tension, and alignment', 'YES', 'QIP-002, PRO-001'],
        ['COMP-003', 'Safety Procedures', 'Quy trình an toàn', '안전 절차', 'SAFETY', 'Understanding and application of workplace safety rules, emergency procedures, and PPE requirements', 'YES', 'QIP-003'],
        ['COMP-004', 'Team Leadership', 'Lãnh đạo nhóm', '팀 리더십', 'LEADERSHIP', 'Ability to guide, motivate, and manage production teams effectively while maintaining quality standards', 'NO', 'QIP-010'],
        ['COMP-005', 'SPC & Data Analysis', 'Phân tích SPC & dữ liệu', 'SPC 및 데이터 분석', 'PROCESS', 'Statistical Process Control knowledge, ability to read control charts, identify trends, and take corrective actions', 'YES', 'QIP-007, QIP-008'],
    ]

    for r, ex in enumerate(examples, 3):
        for c, val in enumerate(ex, 1):
            ws.cell(row=r, column=c, value=val)
        style_row(ws, r, len(headers), EXAMPLE_FILL)

    dv_cat = DataValidation(type='list', formula1='"TECHNICAL,QUALITY,SAFETY,LEADERSHIP,COMMUNICATION,PROCESS"', allow_blank=False)
    ws.add_data_validation(dv_cat)
    dv_cat.add('E3:E500')

    dv_core = DataValidation(type='list', formula1='"YES,NO"', allow_blank=False)
    ws.add_data_validation(dv_core)
    dv_core.add('G3:G500')

    add_legend(ws, 3 + len(examples), len(headers))
    ws.freeze_panes = 'A3'


# ═══════════════════════════════════════════════════════════════════════════════
# 4. TQC TEAMS SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def create_tqc_teams_sheet(wb):
    ws = wb.create_sheet('3. TQC Teams')

    headers = [
        ('team_name *', 20),
        ('team_name_vn', 20),
        ('team_name_kr', 20),
        ('factory', 15),
        ('line', 10),
    ]

    notes = [
        'Team name\n(English)',
        'Vietnamese\nname',
        'Korean\nname',
        'Factory/building\nassignment',
        'Production\nline',
    ]

    for i, (h, w) in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws, 1, len(headers))

    for i, n in enumerate(notes, 1):
        ws.cell(row=2, column=i, value=n).font = NOTE_FONT
        ws.cell(row=2, column=i).alignment = CENTER
        ws.cell(row=2, column=i).border = THIN_BORDER

    examples = [
        ['Team Alpha', 'Đội Alpha', '알파팀', 'BUILDING_A', 'L1'],
        ['Team Beta', 'Đội Beta', '베타팀', 'BUILDING_B', 'L2'],
        ['Team Gamma', 'Đội Gamma', '감마팀', 'BUILDING_C', 'L3'],
    ]

    for r, ex in enumerate(examples, 3):
        for c, val in enumerate(ex, 1):
            ws.cell(row=r, column=c, value=val)
        style_row(ws, r, len(headers), EXAMPLE_FILL)

    add_legend(ws, 3 + len(examples), len(headers))
    ws.freeze_panes = 'A3'


# ═══════════════════════════════════════════════════════════════════════════════
# 5. TRAINING COSTS SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def create_training_costs_sheet(wb):
    ws = wb.create_sheet('4. Training Costs')

    headers = [
        ('program_code *', 18),
        ('period *\n(YYYY-MM)', 15),
        ('trainer_cost\n(VND)', 18),
        ('material_cost\n(VND)', 18),
        ('facility_cost\n(VND)', 18),
        ('other_cost\n(VND)', 18),
        ('notes', 30),
    ]

    notes = [
        'Existing program\ncode',
        'Cost period\n(month)',
        'Trainer salary\nor fee',
        'Books, handouts,\nsupplies',
        'Room rental,\nequipment',
        'Travel, meals,\netc.',
        'Cost notes',
    ]

    for i, (h, w) in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws, 1, len(headers))

    for i, n in enumerate(notes, 1):
        ws.cell(row=2, column=i, value=n).font = NOTE_FONT
        ws.cell(row=2, column=i).alignment = CENTER
        ws.cell(row=2, column=i).border = THIN_BORDER

    examples = [
        ['QIP-001', '2026-01', 5000000, 1200000, 500000, 300000, 'Monthly QIP basic training'],
        ['QIP-002', '2026-01', 8000000, 2000000, 500000, 0, 'Advanced quality control'],
        ['PRO-001', '2026-02', 3000000, 800000, 0, 150000, 'On-the-job training'],
    ]

    for r, ex in enumerate(examples, 3):
        for c, val in enumerate(ex, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if isinstance(val, int):
                cell.number_format = '#,##0'
        style_row(ws, r, len(headers), EXAMPLE_FILL)

    add_legend(ws, 3 + len(examples), len(headers))
    ws.freeze_panes = 'A3'


# ═══════════════════════════════════════════════════════════════════════════════
# INSTRUCTIONS SHEET (Overview)
# ═══════════════════════════════════════════════════════════════════════════════

def create_instructions_sheet(wb):
    ws = wb.create_sheet('Instructions', 0)  # First sheet
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 80

    title_font = Font(name='Calibri', bold=True, size=16, color='1E40AF')
    subtitle_font = Font(name='Calibri', bold=True, size=13, color='1E40AF')
    body_font = Font(name='Calibri', size=11)
    bold_font = Font(name='Calibri', bold=True, size=11)

    r = 2
    ws.cell(row=r, column=2, value='Q-TRAIN System — Master Data Registration Template').font = title_font
    r += 1
    ws.cell(row=r, column=2, value='HWK Vietnam (화승비나) QIP Training Management System').font = Font(size=11, color='6B7280')
    r += 2

    ws.cell(row=r, column=2, value='OVERVIEW').font = subtitle_font
    r += 1
    ws.cell(row=r, column=2, value='This workbook contains 4 data templates that need to be filled to set up the Q-TRAIN system.').font = body_font
    r += 1
    ws.cell(row=r, column=2, value='Employees and Training Programs are already in the system and do NOT need to be included here.').font = bold_font
    r += 2

    ws.cell(row=r, column=2, value='SHEETS IN THIS WORKBOOK').font = subtitle_font
    r += 1

    sheets_info = [
        ('1. Trainers', 'Internal and external trainers with specializations', 'HIGH - Needed before creating sessions'),
        ('2. Competencies', 'Competency framework with multilingual names', 'HIGH - Needed for skill gap analysis'),
        ('3. TQC Teams', 'New employee training team assignments', 'MEDIUM - For new TQC trainee management'),
        ('4. Training Costs', 'Historical and planned training cost data', 'MEDIUM - For ROI analysis & budgeting'),
    ]

    for sheet_name, desc, priority in sheets_info:
        ws.cell(row=r, column=2, value=f'{sheet_name}').font = bold_font
        r += 1
        ws.cell(row=r, column=2, value=f'   {desc}').font = body_font
        r += 1
        ws.cell(row=r, column=2, value=f'   Priority: {priority}').font = Font(size=10, italic=True, color='DC2626' if 'CRITICAL' in priority else '6B7280')
        r += 2

    ws.cell(row=r, column=2, value='HOW TO FILL').font = subtitle_font
    r += 1
    instructions = [
        '1. Each sheet has yellow example rows — use them as reference, then delete or overwrite them.',
        '2. Fields marked with * (asterisk) are REQUIRED. All other fields are optional.',
        '3. Dropdown lists are provided for standardized fields (department, position, etc.).',
        '4. Date format: YYYY-MM-DD (e.g., 2024-03-15)',
        '5. For comma-separated fields, use commas between values (e.g., "QIP-001, QIP-002").',
        '6. Do NOT modify the header rows (rows 1-2).',
        '7. Start entering your data from row 3 (replace example data).',
    ]
    for inst in instructions:
        ws.cell(row=r, column=2, value=inst).font = body_font
        r += 1

    r += 1
    ws.cell(row=r, column=2, value='DEADLINE & SUBMISSION').font = subtitle_font
    r += 1
    ws.cell(row=r, column=2, value='Please complete and return this file by: [TO BE CONFIRMED]').font = bold_font
    r += 1
    ws.cell(row=r, column=2, value='Send completed file to: ksmoon@hsvina.com').font = body_font
    r += 2
    ws.cell(row=r, column=2, value='Questions? Contact the Q-TRAIN system administrator.').font = Font(size=10, color='6B7280')

    ws.sheet_properties.tabColor = '1E40AF'


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    create_instructions_sheet(wb)
    create_trainers_sheet(wb)
    create_competencies_sheet(wb)
    create_tqc_teams_sheet(wb)
    create_training_costs_sheet(wb)

    # Set tab colors
    wb['1. Trainers'].sheet_properties.tabColor = 'F59E0B'
    wb['2. Competencies'].sheet_properties.tabColor = '10B981'
    wb['3. TQC Teams'].sheet_properties.tabColor = '6366F1'
    wb['4. Training Costs'].sheet_properties.tabColor = '8B5CF6'

    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(output_dir, 'Q-TRAIN_Master_Data_Template.xlsx')
    wb.save(output_path)
    print(f'Template saved to: {output_path}')
    return output_path


if __name__ == '__main__':
    main()
